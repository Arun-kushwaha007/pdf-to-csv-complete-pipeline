import os
import re
import pandas as pd
from google.cloud import documentai_v1 as documentai
from google.api_core.client_options import ClientOptions
from typing import List, Dict, Optional
import logging
from datetime import datetime
from nameparser import HumanName
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
import unicodedata

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

class WorkingDocumentProcessor:
    def __init__(self, project_id: str, location: str):
        self.project_id = project_id
        self.location = location
        
        # For 'us' region, use the global endpoint; for other regions, use regional endpoint
        if location == 'us':
            opts = ClientOptions(api_endpoint="documentai.googleapis.com")
        else:
            opts = ClientOptions(api_endpoint=f"{location}-documentai.googleapis.com")
        self.client = documentai.DocumentProcessorServiceClient(client_options=opts)
    
    def process_document(self, file_path: str, custom_processor_id: str) -> Dict:
        logger.info(f"Processing document: {os.path.basename(file_path)}")

        document = self._call_custom_extractor(file_path, custom_processor_id)

        entities = self._extract_entities_simple(document)
        logger.info(f"Found {len(entities)} entities")
        
        if len(entities) == 0:
            logger.error("NO ENTITIES FOUND - Check processor ID and document format")
            return {'raw_records': [], 'filtered_records': []}
        
        records = self._simple_grouping(entities)
        logger.info(f"Grouped into {len(records)} records")
        
        # Clean and validate records (raw data)
        clean_records = self._clean_and_validate(records)
        logger.info(f"Clean records: {len(clean_records)}")
        
        # Apply deduplication (filtered data)
        filtered_records = self._deduplicate_records(clean_records)
        logger.info(f"Filtered records: {len(filtered_records)}")
        
        return {
            'raw_records': clean_records,
            'filtered_records': filtered_records
        }
    
    def _call_custom_extractor(self, file_path: str, processor_id: str):
        processor_name = self.client.processor_path(
            self.project_id, self.location, processor_id
        )
        
        with open(file_path, "rb") as f:
            content = f.read()
        
        request = documentai.ProcessRequest(
            name=processor_name,
            raw_document=documentai.RawDocument(
                content=content,
                mime_type="application/pdf"
            )
        )
        
        result = self.client.process_document(request=request)
        return result.document
    
    def _extract_entities_simple(self, document):
        entities = []
        
        for entity in document.entities:
            cleaned_value = self._clean_text(entity.mention_text)
            if cleaned_value:  # Only add non-empty cleaned values
                entities.append({
                    'type': entity.type_.lower().strip(),
                    'value': cleaned_value
                })
        
        return entities
    
    def _clean_text(self, text: str) -> str:
        """Remove emojis, special characters, and normalize text"""
        if not text:
            return ""
        
        # Remove emojis and special characters
        text = unicodedata.normalize('NFKD', text)
        text = ''.join(c for c in text if not unicodedata.combining(c))
        
        # Remove emojis and other non-printable characters
        emoji_pattern = re.compile(
            "["
            "\U0001F600-\U0001F64F"  # emoticons
            "\U0001F300-\U0001F5FF"  # symbols & pictographs
            "\U0001F680-\U0001F6FF"  # transport & map symbols
            "\U0001F1E0-\U0001F1FF"  # flags (iOS)
            "\U00002702-\U000027B0"  # dingbats
            "\U000024C2-\U0001F251"  # enclosed characters
            "]+", flags=re.UNICODE)
        
        text = emoji_pattern.sub('', text)
        
        # Remove other special characters but keep basic punctuation
        text = re.sub(r'[^\w\s\-.,@()/]', '', text)
        
        # Clean up extra whitespace
        text = ' '.join(text.split())
        
        return text.strip()
    
    def _simple_grouping(self, entities: List[Dict]) -> List[Dict]:
        records = []
        
        # Extract all field types - match exact processor field names
        names = [e['value'] for e in entities if e['type'].lower() == 'name']
        mobiles = [e['value'] for e in entities if e['type'].lower() == 'mobile']
        addresses = [e['value'] for e in entities if e['type'].lower() == 'address']
        emails = [e['value'] for e in entities if e['type'].lower() == 'email']
        landlines = [e['value'] for e in entities if e['type'].lower() == 'landline']
        dobs = [e['value'] for e in entities if e['type'].lower() == 'dateofbirth']
        last_seen = [e['value'] for e in entities if e['type'].lower() == 'lastseen']
        
        logger.info(f"Found: {len(names)} names, {len(mobiles)} mobiles, {len(addresses)} addresses, {len(emails)} emails, {len(landlines)} landlines, {len(dobs)} DOBs, {len(last_seen)} last seen")
        
        # Debug: Show sample entities for troubleshooting
        if names:
            logger.info(f"Sample names: {names[:3]}")
        if addresses:
            logger.info(f"Sample addresses: {addresses[:3]}")
        
        # Use the maximum count to ensure we capture all records
        field_counts = [len(names), len(mobiles), len(addresses), len(emails), len(landlines), len(dobs), len(last_seen)]
        max_count = max(field_counts) if any(field_counts) else 0
        
        for i in range(max_count):
            record = {}
            
            if i < len(names):
                record['name'] = names[i]
            if i < len(mobiles):
                record['mobile'] = mobiles[i]
            if i < len(addresses):
                record['address'] = addresses[i]
            if i < len(emails):
                record['email'] = emails[i]
            if i < len(landlines):
                record['landline'] = landlines[i]
            if i < len(dobs):
                record['date_of_birth'] = dobs[i]
            if i < len(last_seen):
                record['last_seen_date'] = last_seen[i]
            
            # Only add records that have at least a name
            if record.get('name'):
                records.append(record)
        
        return records
    
    def _parse_name(self, full_name: str) -> tuple:
        """Parse full name into first and last name - strict validation only.

        Robust fixes:
        - If the field contains separators (comma/semicolon/slash), pick the segment
        with the most Latin letters (prefer the real name when a stray prefix exists).
        - Remove any leading tokens that are "junk" (punctuation-only or contain no
        Latin/alphanumeric characters), e.g. '...', '오,', '***', etc.
        - Keep the rest of your original checks (addresses, numbers, word lists).
        """
        if not full_name or not full_name.strip():
            return "", ""

        name = full_name.strip()

        # If there are separators, choose the best segment (most Latin letters, then length)
        parts = re.split(r'[;,/\\]\s*', name)
        if len(parts) > 1:
            def latin_score(s: str) -> int:
                # Count basic ASCII Latin letters as a fast proxy
                return len(re.findall(r'[A-Za-z]', s))

            best = max(parts, key=lambda s: (latin_score(s), len(s)))
            if latin_score(best) >= 1 or len(best.split()) >= 2:
                name = best.strip()

        # Split into tokens and drop leading junk tokens (punctuation-only or containing no letters/digits)
        tokens = name.split()
        def is_junk_token(tok: str) -> bool:
            # strip common separators and punctuation edges
            stripped = tok.strip(" ,.;:-_()[]{}\"'`")
            # If nothing left after stripping punctuation, it's junk.
            if stripped == "":
                return True
            # If there are no ASCII letters or digits at all, treat as junk (covers foreign glyph tokens)
            if not re.search(r'[A-Za-z0-9]', stripped):
                return True
            # Otherwise it's probably useful
            return False

        while tokens and is_junk_token(tokens[0]):
            tokens.pop(0)

        name = " ".join(tokens).strip()
        if not name:
            return "", ""

        # If it contains numbers, likely not a name
        if re.search(r'\d', name):
            return "", ""

        # Address-word blacklist (whole words)
        address_words = ['street', 'avenue', 'road', 'drive', 'lane', 'court', 'place', 'way',
                        'crescent', 'close', 'terrace', 'parade', 'boulevard', 'gordonvale',
                        'qld', 'nsw', 'vic', 'wa', 'sa', 'tas', 'nt', 'act',]

        name_lower = name.lower()
        for word in address_words:
            if f' {word} ' in f' {name_lower} ' or name_lower.startswith(f'{word} ') or name_lower.endswith(f' {word}'):
                return "", ""

        name_parts = name.split()
        if len(name_parts) < 2:
            return "", ""

        first_name = name_parts[0].strip()
        last_name = " ".join(name_parts[1:]).strip()

        # Reject names containing digits
        if re.search(r'\d', first_name) or re.search(r'\d', last_name):
            return "", ""

        # Ensure first name has at least 2 letters and contains a Latin letter
        if len(re.sub(r'[^A-Za-z]', '', first_name)) < 2:
            return "", ""

        return first_name, last_name


    
    def _clean_phone_number(self, phone: str) -> str:
        """Clean and validate phone number - strict 10 digit validation"""
        if not phone:
            return ""
        
        # Remove all non-digit characters
        digits = re.sub(r'\D', '', phone)
        
        # Only accept exactly 10 digits
        if len(digits) == 10:
            return digits
        
        # Return empty string for invalid phone numbers
        return ""
    
    def _clean_and_validate(self, records: List[Dict]) -> List[Dict]:
        """Clean and validate all records with comprehensive field handling"""
        clean_records = []
        
        for record in records:
            # Extract and clean all fields
            name = record.get('name', '').strip()
            mobile = record.get('mobile', '').strip()
            address = record.get('address', '').strip()
            email = record.get('email', '').strip()
            landline = record.get('landline', '').strip()
            dob = record.get('date_of_birth', '').strip()
            last_seen = record.get('last_seen_date', '').strip()
            
            # Skip records without a name
            if not name:
                continue
            
            # Parse name into first and last name
            first_name, last_name = self._parse_name(name)
            
            # Skip if we can't get valid first and last names
            if not first_name or not last_name:
                continue
            
            # Clean phone numbers
            mobile_clean = self._clean_phone_number(mobile)
            landline_clean = self._clean_phone_number(landline)
            
            # Mobile number is required (exactly 10 digits)
            if not mobile_clean:
                continue
            
            # Basic email validation
            email_clean = ""
            if email and "@" in email and "." in email.split("@")[-1]:
                email_clean = email.lower().strip()
            
            # Address validation - check first 5 characters for numbers
            address_clean = address
            if address:
                address = address.strip()
                # Remove addresses that are too short
                if len(address) < 15:
                    address_clean = ""
                # Check if first 5 characters contain at least one number
                elif not re.search(r'\d', address[:10]):
                    address_clean = ""
            
            # Address is required
            if not address_clean:
                continue
            
            # Create clean record
            clean_record = {
                'first_name': first_name,
                'last_name': last_name,
                'mobile': mobile_clean,
                'landline': landline_clean,
                'address': address_clean,
                'email': email_clean,
                'date_of_birth': dob,
                'last_seen_date': last_seen
            }
            
            clean_records.append(clean_record)
        
        return clean_records
    
    def _deduplicate_records(self, records: List[Dict]) -> List[Dict]:
        """Remove duplicates based on mobile numbers, keeping the record with most data"""
        phone_to_record = {}
        
        for record in records:
            mobile = record.get('mobile', '')
            
            # Only process records with mobile numbers
            if not mobile:
                continue
            
            # If this mobile number is new, store the record
            if mobile not in phone_to_record:
                phone_to_record[mobile] = record
            else:
                # Compare data completeness and keep the one with more data
                existing_record = phone_to_record[mobile]
                existing_data_count = self._count_data_fields(existing_record)
                current_data_count = self._count_data_fields(record)
                
                if current_data_count > existing_data_count:
                    phone_to_record[mobile] = record
        
        unique_records = list(phone_to_record.values())
        logger.info(f"Deduplication: {len(records)} -> {len(unique_records)} records")
        return unique_records
    
    def _count_data_fields(self, record: Dict) -> int:
        """Count non-empty data fields in a record"""
        count = 0
        for key, value in record.items():
            if value and str(value).strip():
                count += 1
        return count
    
    def save_csv(self, records: List[Dict], output_path: str):
        """Save records to CSV format"""
        if not records:
            logger.warning("No records to save")
            return
        
        df = pd.DataFrame(records)
        # Ensure all expected columns are present
        expected_columns = ['first_name', 'last_name', 'mobile', 'landline', 'address', 'email', 'date_of_birth', 'last_seen_date']
        for col in expected_columns:
            if col not in df.columns:
                df[col] = ''
        
        # Reorder columns to match expected format
        df = df[expected_columns]
        df.to_csv(output_path, index=False)
        logger.info(f"Saved {len(records)} records to {output_path}")
    
    def save_excel(self, records: List[Dict], output_path: str):
        """Save records to Excel format"""
        if not records:
            logger.warning("No records to save")
            return
        
        df = pd.DataFrame(records)
        # Ensure all expected columns are present
        expected_columns = ['first_name', 'last_name', 'mobile', 'landline', 'address', 'email', 'date_of_birth', 'last_seen_date']
        for col in expected_columns:
            if col not in df.columns:
                df[col] = ''
        
        # Reorder columns
        df = df[expected_columns]
        
        # Create Excel workbook with formatting
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Extracted Data"
        
        # Add headers with formatting
        headers = expected_columns
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = openpyxl.styles.Font(bold=True)
            cell.fill = openpyxl.styles.PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Add data
        for row_num, record in enumerate(records, 2):
            for col_num, col_name in enumerate(expected_columns, 1):
                ws.cell(row=row_num, column=col_num, value=record.get(col_name, ''))
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(output_path)
        logger.info(f"Saved {len(records)} records to {output_path}")
    
    def save_batch_excel(self, batch_data: List[Dict], output_path: str, batch_size: int = 25):
        """Save multiple batches to Excel with separate sheets"""
        if not batch_data:
            logger.warning("No batch data to save")
            return
        
        wb = openpyxl.Workbook()
        # Remove default sheet
        wb.remove(wb.active)
        
        expected_columns = ['first_name', 'last_name', 'mobile', 'landline', 'address', 'email', 'date_of_birth', 'last_seen_date']
        
        # Create sheets for each batch
        for batch_idx, batch in enumerate(batch_data):
            sheet_name = f"Batch_{batch_idx + 1}"
            ws = wb.create_sheet(title=sheet_name)
            
            # Add headers
            for col_num, header in enumerate(expected_columns, 1):
                cell = ws.cell(row=1, column=col_num, value=header)
                cell.font = openpyxl.styles.Font(bold=True)
                cell.fill = openpyxl.styles.PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            
            # Add data
            records = batch.get('records', [])
            for row_num, record in enumerate(records, 2):
                for col_num, col_name in enumerate(expected_columns, 1):
                    ws.cell(row=row_num, column=col_num, value=record.get(col_name, ''))
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(output_path)
        logger.info(f"Saved {len(batch_data)} batches to {output_path}")
    
    def group_pdfs_by_count(self, pdf_files: List[str], group_size: int = 25) -> List[List[str]]:
        """Group PDF files by count instead of records"""
        groups = []
        for i in range(0, len(pdf_files), group_size):
            group = pdf_files[i:i + group_size]
            groups.append(group)
        return groups
    
    def process_pdf_group(self, pdf_group: List[str], processor_id: str) -> Dict:
        """Process a group of PDFs and return combined results"""
        all_raw_records = []
        all_filtered_records = []
        group_results = []
        
        for pdf_path in pdf_group:
            try:
                result = self.process_document(pdf_path, processor_id)
                raw_records = result.get('raw_records', [])
                filtered_records = result.get('filtered_records', [])
                
                # Add source file information to each record
                for record in raw_records:
                    record['source_file'] = os.path.basename(pdf_path)
                for record in filtered_records:
                    record['source_file'] = os.path.basename(pdf_path)
                
                all_raw_records.extend(raw_records)
                all_filtered_records.extend(filtered_records)
                
                group_results.append({
                    'file': os.path.basename(pdf_path),
                    'status': 'success',
                    'raw_count': len(raw_records),
                    'filtered_count': len(filtered_records)
                })
                
            except Exception as e:
                group_results.append({
                    'file': os.path.basename(pdf_path),
                    'status': 'error',
                    'error': str(e),
                    'raw_count': 0,
                    'filtered_count': 0
                })
                logger.error(f"Error processing {pdf_path}: {e}")
        
        return {
            'raw_records': all_raw_records,
            'filtered_records': all_filtered_records,
            'group_results': group_results,
            'total_files': len(pdf_group),
            'successful_files': len([r for r in group_results if r['status'] == 'success'])
        }
